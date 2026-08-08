from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from opportunity_agent.models.report import OpportunityReport

logger = logging.getLogger(__name__)


class ExcelReportExporter:
    """
    Generates styled Excel reports from OpportunityReport instances using openpyxl.
    """

    HEADERS = [
        "Match Score (%)",
        "Title",
        "Organization",
        "Country",
        "Job Type",
        "Deadline",
        "Key Strengths",
        "Missing Requirements",
        "Link",
        "Reasoning",
    ]

    def export(
        self,
        report: OpportunityReport,
        output_path: str | Path | None = None,
    ) -> Path:
        """
        Exports the OpportunityReport to an Excel file with conditional formatting,
        freeze panes, auto-fit column widths, and clickable links.

        Args:
            report: The opportunity report to export.
            output_path: Optional explicit file output path. If None, auto-generates
                         a filename formatted as `jobs_YYYY_MM_DD.xlsx`.

        Returns:
            Path: The resolved file path of the saved Excel report.
        """
        if output_path is None:
            date_str = report.generated_at.strftime("%Y_%m_%d")
            output_path = Path(f"jobs_{date_str}.xlsx")
        else:
            output_path = Path(output_path)

        # Ensure parent directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Matched Opportunities"

        # 1. Write Header Row
        ws.append(self.HEADERS)
        self._format_header_row(ws)

        # 2. Populate Data Rows
        for match in report.results:
            job = match.job
            strengths_str = ", ".join(match.strengths) if match.strengths else "N/A"
            missing_str = (
                ", ".join(match.missing_requirements)
                if match.missing_requirements
                else "None"
            )
            deadline_str = job.deadline.strftime("%Y-%m-%d") if job.deadline else "N/A"
            job_type_str = (
                job.job_type.value
                if hasattr(job.job_type, "value")
                else str(job.job_type or "N/A")
            )

            row_data = [
                round(match.overall_score, 1),
                job.title,
                job.organization,
                job.country or "N/A",
                job_type_str,
                deadline_str,
                strengths_str,
                missing_str,
                "Link" if job.url else "N/A",
                match.reasoning or "",
            ]
            ws.append(row_data)

            # Set hyperlinking on the 'Link' column (Column 9 / I)
            row_idx = ws.max_row
            if job.url:
                cell = ws.cell(row=row_idx, column=9)
                cell.hyperlink = job.url
                cell.font = Font(name="Calibri", color="0563C1", underline="single")

        # 3. Apply Conditional Formatting to Match Score Column (Column A)
        self._apply_score_conditional_formatting(ws)

        # 4. Set Freeze Panes (Freeze Header Row)
        ws.freeze_panes = "A2"

        # 5. Enable Auto-Filter across all columns
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

        # 6. Apply Auto Column Widths & Alignment Styles
        self._format_grid_and_widths(ws)

        wb.save(output_path)
        logger.info(
            f"Successfully exported {len(report.results)} matches to {output_path}"
        )
        return output_path

    def _format_header_row(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        Styles the top header row with dark blue fill, bold white text, and borders.
        """
        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="medium", color="1F4E78"),
        )

        ws.row_dimensions[1].height = 28

        for col_num in range(1, len(self.HEADERS) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

    def _apply_score_conditional_formatting(
        self, ws: openpyxl.worksheet.worksheet.Worksheet
    ) -> None:
        """
        Applies conditional formatting to Column A (Score):
        - Score >= 90: Light Green fill
        - 75 <= Score < 90: Light Yellow fill
        - Score < 75: Light Red fill
        """
        if ws.max_row <= 1:
            return

        score_range = f"A2:A{ws.max_row}"

        # Fills and Fonts for score tiers
        green_fill = PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        )
        green_font = Font(color="375623", bold=True)

        yellow_fill = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )
        yellow_font = Font(color="7F6000", bold=True)

        red_fill = PatternFill(
            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
        )
        red_font = Font(color="C65911", bold=True)

        # Rules
        rule_green = CellIsRule(
            operator="greaterThanOrEqual",
            formula=["90"],
            fill=green_fill,
            font=green_font,
        )
        rule_yellow = CellIsRule(
            operator="between",
            formula=["75", "89.99"],
            fill=yellow_fill,
            font=yellow_font,
        )
        rule_red = CellIsRule(
            operator="lessThan", formula=["75"], fill=red_fill, font=red_font
        )

        ws.conditional_formatting.add(score_range, rule_green)
        ws.conditional_formatting.add(score_range, rule_yellow)
        ws.conditional_formatting.add(score_range, rule_red)

    def _format_grid_and_widths(
        self, ws: openpyxl.worksheet.worksheet.Worksheet
    ) -> None:
        """Adjusts column widths based on cell contents and formats cell alignments."""
        thin_border = Border(
            left=Side(style="thin", color="E0E0E0"),
            right=Side(style="thin", color="E0E0E0"),
            top=Side(style="thin", color="E0E0E0"),
            bottom=Side(style="thin", color="E0E0E0"),
        )

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                # Apply subtle borders to data rows
                if cell.row > 1:
                    cell.border = thin_border
                    # Center align specific numeric / short text columns
                    if cell.column in (
                        1,
                        4,
                        5,
                        6,
                        9,
                    ):  # Score, Country, Type, Deadline, Link
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                    else:
                        cell.alignment = Alignment(vertical="center", wrap_text=True)

                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)

            # Cap column widths between 12 and 50 characters for visual balance
            adjusted_width = max(max_len + 3, 12)
            if adjusted_width > 50:
                adjusted_width = 50

            ws.column_dimensions[col_letter].width = adjusted_width
