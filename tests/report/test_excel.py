from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest

from opportunity_agent.models.candidate import CandidateProfile
from opportunity_agent.models.job import JobPosting, JobType
from opportunity_agent.models.match import MatchResult
from opportunity_agent.models.report import OpportunityReport
from opportunity_agent.report.excel import ExcelReportExporter


@pytest.fixture
def sample_report() -> OpportunityReport:
    candidate = CandidateProfile(name="Alice Scientist")

    job1 = JobPosting(
        id="1",
        title="Postdoc in Quantum Physics",
        organization="Max Planck Institute",
        country="Germany",
        job_type=JobType.POSTDOC,
        deadline=date(2026, 12, 31),
        url="https://example.com/job1",
    )
    match1 = MatchResult(
        job=job1,
        overall_score=92.5,
        skill_score=90.0,
        education_score=95.0,
        experience_score=90.0,
        research_score=95.0,
        publication_score=90.0,
        strengths=["Strong quantum background"],
        missing_requirements=["CUDA"],
        reasoning="Excellent research fit.",
    )

    job2 = JobPosting(
        id="2",
        title="Data Scientist",
        organization="CERN",
        country="Switzerland",
        job_type=JobType.RESEARCH_ENGINEER,
        deadline=date(2026, 9, 15),
        url="https://example.com/job2",
    )
    match2 = MatchResult(
        job=job2,
        overall_score=78.0,
        skill_score=75.0,
        education_score=80.0,
        experience_score=75.0,
        research_score=80.0,
        publication_score=80.0,
        strengths=["Python proficiency"],
        missing_requirements=["C++"],
        reasoning="Good fit.",
    )

    job3 = JobPosting(
        id="3",
        title="Junior Analyst",
        organization="Tech Corp",
        country="France",
        job_type=JobType.OTHER,
        deadline=None,
        url="",
    )
    match3 = MatchResult(
        job=job3,
        overall_score=60.0,
        skill_score=60.0,
        education_score=60.0,
        experience_score=60.0,
        research_score=60.0,
        publication_score=60.0,
        strengths=[],
        missing_requirements=["PhD Degree"],
        reasoning="Low match score.",
    )

    return OpportunityReport(
        candidate=candidate,
        results=[match1, match2, match3],
        generated_at=datetime(2026, 8, 9, 12, 0, 0),
    )


def test_excel_export_creates_file(
    tmp_path: Path, sample_report: OpportunityReport
) -> None:
    """Verify that export generates an Excel file with expected rows and structure."""
    output_file = tmp_path / "test_report.xlsx"
    exporter = ExcelReportExporter()

    result_path = exporter.export(sample_report, output_path=output_file)

    assert result_path.exists()

    wb = openpyxl.load_workbook(result_path)
    sheet = wb.active

    assert sheet.title == "Matched Opportunities"
    assert sheet.max_row == 4  # 1 Header row + 3 Data rows
    assert sheet.max_column == 10

    # Assert Header Titles
    assert sheet.cell(row=1, column=1).value == "Match Score (%)"
    assert sheet.cell(row=1, column=2).value == "Title"

    # Assert Row 2 Data
    assert sheet.cell(row=2, column=1).value == 92.5
    assert sheet.cell(row=2, column=2).value == "Postdoc in Quantum Physics"
    assert sheet.cell(row=2, column=4).value == "Germany"

    # Assert Hyperlink on Row 2 Column 9 (Link)
    link_cell = sheet.cell(row=2, column=9)
    assert link_cell.value == "Link"
    assert link_cell.hyperlink.target == "https://example.com/job1"

    # Assert Freeze Panes
    assert sheet.freeze_panes == "A2"

    # Assert Auto-Filter
    assert sheet.auto_filter.ref is not None


def test_excel_export_auto_filename(
    tmp_path: Path, sample_report: OpportunityReport, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Verify default filename generation when no path is provided."""
    monkeypatch.chdir(tmp_path)
    exporter = ExcelReportExporter()

    result_path = exporter.export(sample_report)

    assert result_path.name == "jobs_2026_08_09.xlsx"
    assert result_path.exists()
